import openpyxl
workbook = openpyxl.load_workbook("C:Users\\Madhu Kumar PSI-3442\\Desktop\\Practice.xlsx")
sheet = workbook["Sheet1"]

total_rows = sheet.max_row
total_Cols = sheet.max_column

print(total_Cols,total_rows)
print(sheet.cell(3,1).value)

for r in range(1,total_rows+1):
    for c in range(1,total_Cols+1):
        print(sheet.cell(r,c).value,end="        ")
    print()

#write data into excel file




row_data = ["maddy@test.com","pinky@test.com","monky@test.com"]
col_data = ["mad@1234","pky@1234","mnky@test.com"]

for i in range(total_rows+1,total_rows+len(row_data)+1):
    for j in range(1,total_Cols+1):
        if j==1:
            sheet.cell(i, j).value =row_data[i-total_rows-1]
        if j==2:
            sheet.cell(i, j).value = col_data[i-total_rows-1]

workbook.save("C:Users\\Madhu Kumar PSI-3442\\Desktop\\Practice.xlsx")

