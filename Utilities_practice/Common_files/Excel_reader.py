import openpyxl

def Extract_Excel_data(Path,Sheet_Name):
    Final_list = []
    work_book = openpyxl.load_workbook(Path)
    sheet = work_book[Sheet_Name]
    total_rows = sheet.max_row
    total_cols = sheet.max_column

    for r in range(2,total_rows+1):
        row_list = []
        for c in range(1,total_cols+1):
            row_list.append(sheet.cell(r,c).value)
        Final_list.append(row_list)

    return Final_list
    # return [
    #     ["madhu@test.com", "mad@1234"],
    #     ["pinky@test.com", "plm@143"]
    # ]
